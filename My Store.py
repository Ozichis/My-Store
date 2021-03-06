#Modules
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QGroupBox, QHBoxLayout, QVBoxLayout, QPlainTextEdit, QGridLayout, QCheckBox, QMainWindow, QInputDialog, QMessageBox, QComboBox, QSpinBox, QLineEdit, QCompleter, QAction, QScrollArea, QAbstractScrollArea
from tkinter import Tk, Entry
import wikipedia
import datetime
import sys
import random
import openpyxl as xl
from PyQt5.QtGui import QIcon, QFont, QPixmap
import requests
from PyQt5.QtCore import QRect, QSize, Qt, QRect
import random
r = random.randint(1, 10000)
login = False
e = 0
apples = [0]
apples_to_cart = [0]
grapes = 0
grapes_to_cart = 0
bananas = 0
bananas_to_cart = 0
cars = 0
fast_cars = 0
phones = 0
laptops = 0

#Window

class Window(QWidget):
	def __init__(self):
		super().__init__()
		self.create_ui_()
		
	def create_ui_(self):
		self.setWindowTitle('My Store')
		self.setGeometry(100, 100, 800, 300)
		self.load()
		self.show()
		
	def load(self):

		#raise Exception(dir(QAbstractScrollArea))
		#QScrollArea.setHorizontalScrollBar(QAbstractScrollArea.scrollBarWidgets)
		#p = Entry(self)
		#p.pack()
		self.dty = 600
		self.amoun = 2500
		self.gpbox = QGroupBox("Welcome to my Store, here are things available to do")
		self.gpbox.setFont(QFont('Times New Roman', 20))
		self.gpbox.setGeometry(100, 100, 70, 20)
		grid = QGridLayout()
		self.vbox = QVBoxLayout()
		
#		self.s1 = QScrollBar()
#		self.hbox = QHBoxLayout()
#		self.s1.setMaximum(100)
#		self.hbox.addWidget(self.s1)
		
		self.btn = QPushButton('Buy Fruits', self)
		self.btn.setStyleSheet('background-color:red')
		self.btn.clicked.connect(self.buy)
		self.btn.setIconSize(QSize(40,50))
		grid.addWidget(self.btn, 0,0)
		
		self.btn1 = QPushButton('Chat with us', self)
		self.btn1.clicked.connect(self.chat)
		self.btn1.setStyleSheet('background-color:blue')
		self.btn1.setIconSize(QSize(40,50))
		grid.addWidget(self.btn1, 0,1)
		
		self.btn2 = QPushButton ('Check login details', self)
		self.btn2.setStyleSheet('background-color:green')
		self.btn2.clicked.connect(self.detail)
		self.btn2.setIconSize(QSize(40,50))
		grid.addWidget(self.btn2, 1,0)
		
		self.btn9000 = QPushButton ("Check balance details\n--If you don't have..", self)
		self.btn9000.setStyleSheet('background-color:green')
		self.btn9000.clicked.connect(self.detail2)
		self.btn9000.setIconSize(QSize(40,50))
		grid.addWidget(self.btn9000, 4,1)
		
		self.btn90001 = QPushButton ("Gain free prize\n--If connected", self)
		self.btn90001.setStyleSheet('background-color:red')
		self.btn90001.clicked.connect(self.gain)
		self.btn90001.setIconSize(QSize(40,50))
		grid.addWidget(self.btn90001, 5,0)
		
		self.btn3 = QPushButton ('Sign-up', self)
		self.btn3.setStyleSheet('background-color:pink')
		self.btn3.clicked.connect(self.signin)
		self.btn3.setIconSize(QSize(40,50))
		grid.addWidget(self.btn3, 1,1)
		#self.me = QLabel( "Hello", self)
		#self.me.setStyleSheet('border:1px solid black')
		#grid.addWidget(self.me, 4, 1)
		
		
		
		self.btn4 = QPushButton ('Make a plan', self)
		self.btn4.setStyleSheet('background-color:purple')
		self.btn4.clicked.connect(self.makeplan)
		self.btn4.setIconSize(QSize(40,50))
		grid.addWidget(self.btn4, 2,0)
		self.gpbox.setLayout(grid)
		#self.gpbox.setLayout(self.hbox)
		
		self.btn5 = QPushButton ('Send Surveys', self)
		self.btn5.setStyleSheet('background-color:green')
		self.btn5.setIconSize(QSize(40,50))
		grid.addWidget(self.btn5, 2,1)
		
		self.btn6 = QPushButton ('Pay in', self)
		self.btn6.setStyleSheet('background-color:red')
		self.btn6.clicked.connect(self.payin)
		self.btn6.setIconSize(QSize(40,50))
		grid.addWidget(self.btn6, 3,0)
		
		self.btn7 = QPushButton('Open Cart', self)
		self.btn7.setStyleSheet('background-color:green')
		self.btn7.clicked.connect(self.cart_open)
		grid.addWidget(self.btn7, 3, 1)
		
		self.btn8 = QPushButton('Send message', self)
		self.btn8.setStyleSheet('background-color:blue')
		self.btn8.clicked.connect(self.send_a_message)
		grid.addWidget(self.btn8, 4, 0)
		
#		self.hider2 = QPushButton("back")
#		self.hider2.setStyleSheet('background-color:black; color:white')
#		self.hider2.setGeometry(600, 1, 60, 25)
#		self.hider2.showNormal()
#		self.hider2.clicked.connect(self.back)
		#self.vbox.addWidget(file)
		self.vbox.addWidget(self.gpbox)
		#self.vbox.addWidget(self.hbox)
		#self.vbox.addWidget()
		#self.setLayout(self.vbox)
		#self.g = QWidget()
		#self.g.setLayout(self.vbox)
		
		#self.scroll = QScrollArea()
		#self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
		#self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
		#self.scroll.setWidgetResizable(True)
		#self.scroll.setWidget(self)
		#self.vbox.addWidget(self.scroll)
		self.setLayout(self.vbox)
		#self.setCentralWidget(self.vbox)div
		
	#def detailedfd(self, i):
		#if i.text() == "Show Details...":
#		self.msg.setDetailedText( "The details are as follows:\n-This is option is for getting free prizes\n-You can only get a free prize if you are connected \n-There are different prizes you can get\n-Sometimes even though you are connected it can't find a prize")
		
		
	def gain(self):
	   def checkInternetRequests(url='http://www.google.com/', timeout=random.randint(2, 10)):
#	   	self.msg = QMessageBox()
#	   	self.msg.setIcon(QMessageBox.Information)
#	   	self.msg.setText("Would you like to look at the details of this option")
#	   	self.msg.setInformativeText("This is our guide abo.ut this option")
#	   	self.msg.setWindowTitle("Details")
#	   	#self.msg.setDetailedText(f"Will you {dir(self.msg)}")
#	   	self.msg.buttonClicked.connect(self.detailedfd)
#	   	self.msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
#	   	#self.msg.addButton(QMessageBox.Yes)
#	   	retval = self.msg.exec_()
#	   	if type(self.msg).__name__ == QMessageBox.Yes:
#	   		self.msg = QMessageBox()
#	   		self.msg.setIcon(QMessageBox.Information)
#	   		self.msg.setText("Here are the details")
#	   		self.msg.setInformativeText("This is our guide about this option")
#	   		self.msg.setWindowTitle("Details")
#	   		#self.msg.setDetailedText(f"Will you {dir(self.msg)}")self.msg.buttonClicked.connect(self.detailedfd)
#	   		#self.msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
#	   		self.msg.setDetailedText( "The details are as follows:\n-This is option is for getting free prizes\n-You can only get a free prize if you are connected \n-There are different prizes you can get\n-Sometimes even though you are connected it can't find a prize")
#	   		retval = self.msg.exec_()
	   	#QMessageBox.about(self, '&', str(self.msg))
	   	try:
	   	   	r = requests.head(url, timeout=timeout)
	   	   	if random.randint(1, 10) == 1:
	   	   			try:
	   	   				global laptops
	   	   			except:
	   	   				pass
	   	   			laptops += random.randint(1, 10)
	   	   			QMessageBox.information(self, 'Success', f'You now have {laptops} laptops')
	   	   	if random.randint(1, 10) == 2:
	   	   			try:
	   	   				global phones
	   	   			except:
	   	   				pass
	   	   			phones += random.randint(1, 10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {phones} phones")
	   	   	if random.randint(1, 10) == 3:
	   	   			try:
	   	   				global cars
	   	   			except:
	   	   				pass
	   	   			cars += random.randint(1, 10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {cars} cars")
	   	   	if random.randint(1, 10) == 4:
	   	   			try:
	   	   				global fast_cars
	   	   			except:
	   	   				pass
	   	   			fast_cars += random.randint(1, 10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {fast_cars} fast cars")
	   	   	if random.randint(1, 10) == 5:
	   	   			try:
	   	   				global bananas
	   	   			except:
	   	   				pass
	   	   			bananas += random.randint(1, 10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {bananas} bananas")
	   	   	if random.randint(1, 10) == 6:
	   	   			try:
	   	   				global apples
	   	   			except:
	   	   				pass
	   	   			apples[0] += random.randint(1,10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {apples[0]} apples")
	   	   	if random.randint(1,10) == 7:
	   	   			try:
	   	   				global grapes
	   	   			except:
	   	   				pass
	   	   			grapes += random.randint(1,10)
	   	   			QMessageBox.information(self, 'Success', f"You now have {grapes} grapes")
	   	   	if random.randint(1,10) == 8:
	   	   			try:
	   	   				self.amount += random.randint(32, 500)
	   	   				QMessageBox.information(self, 'Success', f"Your balance is now ${self.amount}")
	   	   			except:
	   	   				self.amoun += random.randint(32, 500)
	   	   				QMessageBox.information(self, 'Success', f"Your balance is now ${self.amoun}")
	   	   	if random.randint(1,10) == 9:
	   	   				try:
	   	   					self.data += random.randint(100, 9000)
	   	   					QMessageBox.information(self, 'Success', f'Your data is now {self.data/1200}GB')
	   	   				except:
	   	   					self.dty += random.randint(100, 9000)
	   	   					QMessageBox.information(self, 'Success', f"Your data is now {self.dty/1200}GB")
	   	   	if random.randint(1,10) == 10:
	   	   		   	QMessageBox.information(self, 'Error', f"Can't find a prize right now")
	   	except Exception as ex:
	   	   	QMessageBox.information(self, 'Error', str(type(ex).__name__) + ": " + str(ex))
	       	
	       	
	   checkInternetRequests()
			
		
		
	def detail2(self):
		try:
			
			
			QMessageBox.question(self, 'Error', f'This option is for someone who does not have an account and you have an account with the name {self.check1}', QMessageBox.Ok)
			
			
		except:
			
			
			QMessageBox.about(self, 'Info', f"Balance:{self.amoun}\nData:{self.dty}")
			
			
	def ask_for_message(self):
		message, entered = QInputDialog.getMultiLineText(self, 'Message', 'Enter your message')
		
		
		if str(message) == "":
			QMessageBox.question(self, 'Error', 'Empty message', QMessageBox.Ok)
		if entered:
			response = {
				"hello": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"hello,": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"hi,": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"hi": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"hey": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"hey,": random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]),
				"how are you": random.choice(["Store Bot: Thanks for asking, I am fine", "Store Bot: I am fine, thanks", "Store Bot: I am fine, thank you"]),
				"how are you,": random.choice(["Store Bot: Thanks for asking, I am fine", "Store Bot: I am fine, thanks", "Store Bot: I am fine, thank you"]),
				"what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"hi, what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]), 
				"hello, what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"hey, what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"hey there, what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks \n by writing various code"]),
				"hi,what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"hey there,what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"hello, what is programming": random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing various code"]),
				"what is your name": random.choice(["Store Bot: My Name is Store Bot, there are many things I want to tell you but let me just keep quiet", "Store Bot: My Name is Store Bot", "Store Bot: Ok, you want to know my name my name is Store Bot", "Store Bot: My Name Is Store Bot, that's my name"]),
				"what's your name":random.choice(["Store Bot: My Name is Store Bot, there are many things I want to tell you but let me just keep quiet", "Store Bot: My Name is Store Bot", "Store Bot: Ok, you want to know my name my name is Store Bot", "Store Bot: My Name Is Store Bot, that's my name"]),
				"hello,what is your name": random.choice(["Store Bot: My Name is Store Bot, there are many things I want to tell you but let me just keep quiet", "Store Bot: My Name is Store Bot", "Store Bot: Ok, you want to know my name my name is Store Bot", "Store Bot: My Name Is Store Bot, that's my name"]),
				
				"hello, what is your name": random.choice(["Store Bot: My Name is Store Bot, there are many things I want to tell you but let me just keep quiet", "Store Bot: My Name is Store Bot", "Store Bot: Ok, you want to know my name my name is Store Bot", "Store Bot: My Name Is Store Bot, that's my name"]),
				
	
				"hello,what's your name": random.choice(["Store Bot: My Name is Store Bot, there are many things I want to tell you but let me just keep quiet", "Store Bot: My Name is Store Bot", "Store Bot: Ok, you want to know my name my name is Store Bot", "Store Bot: My Name Is Store Bot, that's my name"]),
			}
			#try:
			self.another_label.setText(self.another_label.text() + "\nYou: " + str(message))
			output = ""
			messages = message.lower().split('.')
			#for word in messages:
				#output += response.get(word, "") + " "
				
			if "hello" in message.lower(): #and "what" and " programming" not in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]))
			
			
			if "hey" and "there" in message.lower():# and "what" and " programming" not in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]))
				
			if "hi" in message.lower() and "my name is" not in message.lower():  #and "what" and " programming" not in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Hi", "Store Bot: Hello", "Store Bot: Hey there"]))
				
			if "what " and "programming" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Programming has to do with writing codes to perform basic tasks", "Store Bot: Programming is giving tasks to the computer in code", "Store Bot: Programming is telling to the computer to do various tasks by writing \nvarious code"]))
				
			if "my name is " in message.lower() and "my name is " != message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Hi " + message.lower().replace('my name is', '').replace('hello', '').replace('cannot add to cart', '').replace( "can't add to cart", "" ).replace('what is programming', "").replace('i', ''),                "Store Bot: Hello "  + message.lower().replace('my name is', '').replace('hello', '').replace( "can't add to cart" , "" ).replace('cannot add to cart', '').replace( "what is programming", "").replace("i", "" ),                                       "Store Bot: Hey there "  + message.lower().replace('my name is', '').replace('hello', '').replace('cannot add to cart', '').replace( "what is programming", "").replace("can't add to cart", "").replace("i", "")]))
				
			if "your" and "name" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: My Name Is Store Bot", "Store Bot: My Name is Store Bot" , "Store Bot: I am Store Bot" ]))
			
			if "thank" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: You Welcome", "Store Bot: I am just happy to help, anyway thank you", "Just happy to help"]))
				
			if "what" and "balance" in message.lower():
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your balance is ${self.amount}", f'Store Bot: I can see that the money in your account is ${self.amount}']))
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your balance is ${self.amoun}", f"Store Bot: The Money in your account is ${self.amoun}"]))
					
					
			if "what" and "money" in message.lower():
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your balance is ${self.amount}", f'Store Bot: I can see that the money in your account is ${self.amount}']))
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your balance is ${self.amoun}", f"Store Bot: The Money in your account is ${self.amoun}"]))
					
			if "your" and "best" and "color" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: Thanks for asking, actually my best color is orange", "Store Bot: Happy to answer, my best color is orange", "Store Bot: My best color is orange", "Store Bot: My best... It is Orange"]))
				
			if "what" and "up" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: There's nothing wrong but I am just looking forward to asking questions to respond to customers", "Store Bot: Customer View", "Store Bot: Just getting ready to answer your questions"]))
				
			if "set" and "orange" in message.lower():
				
				self.setStyleSheet('background-color:orange')
				
				self.another_label.setText(self.another_label.text() + "\n"  +  random.choice([ "Store Bot: Done it"]))
				
				
			if "grateful" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I am so grateful with you also" , "Store Bot: I am also glad to hear your thoughts"]))
				
			
			if "set" and "color" in message.lower():
				
				answer, requested = QInputDialog.getText(self, 'Store', "Store Bot: I am asking you for the color")
				
				self.setStyleSheet("background-color:" + str(answer))
				
			
			
				
			if "login" and "have" in message.lower():
				
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Yes you have an account in this store with the name {self.check1}", 'Store Bot: Of Course, you have an account']))
					
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I have found out that you do not have an account, if you do, you did not sign-up to the account", "Store Bot: I can't find your account, Maybe you don't have an account or you did not sign-up to it", "Store Bot: I cannot find any account"]))
					
					
			if "how are you" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I am fine, thanks for asking", "Store Bot: I am 20% complete", "Store Bot: Thanks for asking me, I am fine, thank you"]))
					
			if "what" and "data" in message.lower():
				
				
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your data is {self.data/1200}GB", f"Store Bot: Your data is {self.data}MB"]))
				
						
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Your data is {self.dty/1200}GB", f"Store Bot: Your data is {self.dty}MB"]))
					
					
			if "sign" and "have" in message.lower():
				
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Yes you have an account in this store with the name {self.check1}", 'Store Bot: Of Course, you have an account']))
					
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I have found out that you do not have an account, if you do, you did not sign-up to the account", "Store Bot: I can't find your account, Maybe you don't have an account or you did not sign-up to it", "Store Bot: I cannot find any account"]))
					
			if "account" and "have" in message.lower():
				
				try:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([f"Store Bot: Yes you have an account in this store with the name {self.check1}", 'Store Bot: Of Course, you have an account']))
					
				except:
					
					self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I have found out that you do not have an account, if you do, you did not sign-up to the account", "Store Bot: I can't find your account, Maybe you don't have an account or you did not sign-up to it", "Store Bot: I cannot find any account"]))
					
			if "cannot add to cart" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I guess that you do not have that amount, to ensure see the price of the amount and check if your balance is up to that amount if it is tell me and we will solve that problem right away", "Store Bot: Maybe you don't have that amount to buy that product", "Store Bot: I think the problem is that you do not have that amount of money"]))
					
			if "can't add to cart" in message.lower():
				self.another_label.setText(self.another_label.text() + "\n" + random.choice([ "Store Bot: I guess that you do not have that amount, to ensure see the price of the amount and check if your balance is up to that amount if it is tell me and we will solve that problem right away", "Store Bot: Maybe you don't have that amount to buy that product", "Store Bot: I think the problem is that you do not have that amount of money"]))
				
			if "&" in message.lower():
				#try:
					#QMessageBox.about(self, 'Correct', str(wikipedia.search( "")))
				#except Exception as ex:
					#self.another_label.setText(str(ex))
				
			
				
			

				
			#elif "my name"
				
			if "how" and "quit" in message.lower():
				
				self.another_label.setText(self.another_label.text() + "\n" + random.choice(["Store Bot: Just click the button below", "Store Bot: By clicking the back button below", "Store Bot: You just need to click the button below"]))
				
			if "quiz" and "play" in message.lower():
				going_first, did = QInputDialog.getText(self, 'Store', "Store Bot: Who wants to start first, if you want to ask, type 'm' if you do not want to ask, type 'y'")
				score = 0
				
				if not did:
					pass
				
				if str(going_first) == "y":
					questss = ['Which came first, cadbury or google, if cadbury type "a" if google type "b" if you want to quit type "q"']
					quests = random.choice(questss)
					if quests == questss[0]:
						answer, did = QInputDialog.getText(self, 'Quiz', 'Store Bot: ' + quests)
						if not did:
							pass
							del questss[0]
							
							
						if str(answer) == "a":
							QMessageBox.about(self, 'Correct', 'Store Bot: You are correct, get ready for the next question')
							del questss[0]
							score += 1
							
						if str(answer) == "b":
							QMessageBox.about(self, 'Wrong', "Cadbury came before google, just get ready for the next question")
						
			#else:
				
				#self.another_label.setText(self.another_label.text() + "\n" + random.choice(['Store Bot: I do not know that question', 'Store Bot: I do not know that']))
				
			try:
				QMessageBox.question(self, 'Store Bot: Wikipedia results', 'Store Bot: ' + str(wikipedia.summary(message.lower())), QMessageBox.Ok)
			except:
				self.another_label.setText(self.another_label.text() + "\nStore Bot: I can't find results from wikipedia")
				
			self.another_label.setText(self.another_label.text() + "\n")
			#except:
				#self.another_label.setText(self.another_label.text() + "\nYou: " + str(message))
#			if "hello".upper:
#				self.another_label.setText(self.another_label.text() + "\nStore Bot: " +  random.choice([ "Hello", "Hi", "Hey there"]))
#			if "HELLO".lower in message:
#				self.another_label.setText(self.another_label.text() + "\nStore Bot: " +  random.choice([ "Hello", "Hi", "Hey there"]))
			#QMessageBox.question(self, 'Yes', self.another_label.text())
			
		
		
	def chat(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		try:
			self.comboboxlabel211.hide()
			self.lkp.hide()
			self.lk21.hide()
			self.b90.hide()
		except:
			pass
			
			
		
		try:
			self.b1.hide()
			self.hider2.hide()
			self.b2.hide()
			self.gpbox2.hide()
			self.b3.hide()
		except:
			pass
		
		try:
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.label1.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.hider.hide()
			self.comboboxlabel2.hide()
			self.btn194.hide()
			self.btn195.hide()
			
		except:
			pass
			
		try:
			self.spin2.hide()
			self.button.hide()
			self.label3.hide()
			self.hier.hide()
			
		except:
			pass
			
		try:
			self.another_label.hide()
			self.sending_button.hide()
			self.hider.hide()
		except:
			pass
		
		try:
			self.another_label = QLabel(self.another_label.text(), self)
		except:
			self.another_label = QLabel( "", self)
		self.another_label.setGeometry(1, 1, 600, 900-25)
		self.another_label.showNormal()
		self.another_label.setStyleSheet('border:1px solid black')
		self.vbox.addWidget(self.another_label)
		self.sending_button = QPushButton("Send", self)
		self.sending_button.showNormal()
		self.sending_button.setGeometry(1, 900, 600, 25)
		self.sending_button.clicked.connect(self.ask_for_message)
		self.sending_button.setStyleSheet("background-color:orange")
		self.vbox.addWidget(self.sending_button)
		self.hider = QPushButton("Back", self)
		self.hider.showNormal()
		self.hider.setGeometry(1, 925, 600, 25)
		self.hider.setStyleSheet("background-color:black")
		self.hider.clicked.connect(self.back)
		self.vbox.addWidget(self.hider)
		
		
	def send_a_message(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		first, done1 = QInputDialog.getText(self, 'Send', 'Enter your name')
		second, done2 = QInputDialog.getItem(self, 'Receiver', 'Who are you sending your message to: ', ['Ozichi', 'Ifechi', 'Akachi'])
		third, done3 = QInputDialog.getMultiLineText(self, 'Message Body', 'Enter your message')
		if str(second) == "Ozichi":
			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/First Customer Information.xlsx")
			sheet = wb["First Customer Information.csv"]
			sheet["B9"] = f"{str(first)}: {str(third)} {datetime.datetime.now()}"

			wb.save("/storage/emulated/0/Lucky bundle cache/First customer Information.xlsx")
			QMessageBox.question(self, 'Success', 'Sent Message Successfully', QMessageBox.Ok)
			
		if str(second) == "Ifechi":
			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/Fourth customer information.xlsx")
			sheet = wb['Sheet1']
			sheet["B9"] = f"{str(first)}: {str(third)} {datetime.datetime.now()}"
			
			wb.save("/storage/emulated/0/Lucky bundle cache/Fourth customer information.xlsx")
			QMessageBox.question(self, 'Success', 'Sent successfully', QMessageBox.Ok)
			
		if str(second) == "Akachi":
			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/Third customer Information.xlsx")
			sheet = wb["Third customer Information.csv"]
			sheet["B9"] = f"{str(first)}: {str(third)} {datetime.datetime.now()}"
			
			wb.save("/storage/emulated/0/Lucky bundle cache/Third customer Information.xlsx")
			QMessageBox.question(self, 'Success', 'Sent message successfully', QMessageBox.Ok)
			
		self.back()
			
			
		
#		self.dk = QLabel("To:", self)
#		#self.dk.showNormal()
#		
#		self.names = ['Ozichi', 'Akachi', 'ifechi']
#		self.completer = QCompleter(self.names)
#		
#		self.btb = QPushButton('Send', self)
#		self.btb.clicked.connect(self.send_the_message)
#		self.btb.setStyleSheet( "background-color:red")
#		self.line_edit = QLineEdit(self)
#		self.line_edisk = QLabel('Message:', self)
#		self.dj = QLabel("From:", self)
#		self.mg = QLineEdit()
#		self.mg.setPlaceholderText('Sender Name')
#		#self.line_edisk.showNormal()
#		#line_edit.Password = True
#		#raise Exception(help(line_edit))
#		#self.line_edit.displayText()
#		self.line_edit.setCompleter(self.completer)
#		self.line_edit.showNormal()
#		self.line_edit.setPlaceholderText('Receiver Name')
#		#self.dk.setText(self.line_edit.text())
#		
#		self.plain_textedit = QPlainTextEdit()
#		self.plain_textedit.setPlaceholderText('Enter your message')
#		#plain_textedit.setReadOnly(True)
#		
#		text = f"Your message: "
#		self.edisk = QLabel( "Message:", self)
#		#self.plain_textedit.appendPlainText(text)
#		self.vbox.addWidget(self.dj)
#		self.vbox.addWidget(self.mg)
#		self.vbox.addWidget(self.dk)
#		self.vbox.addWidget(self.line_edit)
#		self.vbox.addWidget(self.edisk)
#		self.vbox.addWidget(self.plain_textedit)
#		self.vbox.addWidget(self.btb)
		
	def buy_me(self):
		global bananas_to_cart
		self.comboboxlabel21.hide()
		self.lk.hide()
		self.lk2.hide()
		self.b.hide()
		try:
			self.amount -= bananas_to_cart*70
		except:
			self.amoun -= bananas_to_cart*70
			
		self.label1man = QLabel( "❌")
		self.labeldouble = QLabel("No items in your cart")
		if True:
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			bananas_to_cart -= bananas_to_cart
			
	def btny(self):
		global bananas_to_cart
		self.comboboxlabel211.hide()
		self.lkp.hide()
		self.lk21.hide()
		self.b90.hide()
			
		try:
			self.amount -= bananas_to_cart*70
		except:
			self.amoun -= bananas_to_cart*70
		
		bananas_to_cart -= bananas_to_cart
		self.comboboxlabel21.setGeometry(1, 1, 134, 89)
		self.lk.setGeometry(135, 1, 130, 99)
		#self.lk2.setGeometry(1, 1, 130, 99)
		self.b.setGeometry(1, 89, 134, 25)
		
	def buy_me50(self):
		global apples_to_cart
		self.comboboxlabel21.hide()
		self.lk.hide()
		self.lk2.hide()
		self.b.hide()
		try:
			self.amount -= apples_to_cart[0]*50
		except:
			self.amoun -= apples_to_cart[0]*50
			
		self.label1man = QLabel( "❌")
		self.labeldouble = QLabel("No items in your cart")
		if bananas_to_cart == 0 and grapes_to_cart == 0:
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
		apples_to_cart[0] -= apples_to_cart[0]
		


		
				
				
			
			
		
	def cart_open(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		try:
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.label1.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.hider.hide()
			self.comboboxlabel2.hide()
			self.btn194.hide()
			self.btn195.hide()
			
		except:
			pass
			
		try:
			self.spin2.hide()
			self.button.hide()
			self.label3.hide()
			self.hier.hide()
			
		except:
			pass
			
			
		try:
			self.another_label.hide()
		except:
			pass
		try:
			self.b1.hide()
			self.hider2.hide()
			self.b2.hide()
			self.gpbox2.hide()
			self.b3.hide()
		except:
			pass
			
			
		global bananas_to_cart
		global apples_to_cart
			
		if bananas_to_cart == 0 and apples_to_cart[0] == 0 and grapes_to_cart == 0:
			self.label1man = QLabel( "❌", self)
			self.labeldouble = QLabel("No items in your cart", self)
			if True:
				self.label1man.showNormal()
				self.labeldouble.showNormal()
				self.label1man.setGeometry(300, 400, 40, 30)
				self.labeldouble.setGeometry(250, 430, 150, 50)
				self.vbox.addWidget(self.label1man)
				self.vbox.addWidget(self.labeldouble)
				bananas_to_cart -= bananas_to_cart
			
		if bananas_to_cart > 0 and apples_to_cart[0] == 0 and grapes_to_cart == 0:
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-84599807.png')
			self.comboboxlabel21 = QLabel(self)
			self.comboboxlabel21.setPixmap(self.combobox21)
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk = QLabel(str(bananas_to_cart) + " pieces", self)
			self.lk.showNormal()
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2 = QLabel("$" + str(bananas_to_cart*70), self)
			self.lk2.showNormal()
			#self.lk2.setGeometry(135, 0, 130, 99)
			self.comboboxlabel21.showNormal()
			self.b = QPushButton(f"Buy {bananas_to_cart} bananas", self)
			self.b.clicked.connect(self.buy_me)
			self.b.setGeometry(1, 89, 134, 25)
			self.b.setStyleSheet('background-color:green')
			self.b.showNormal()
			try:
				if self.amount < bananas_to_cart*70:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			except:
				if self.amoun < bananas_to_cart*70:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel21)
			self.vbox.addWidget(self.lk)
			self.vbox.addWidget(self.lk2)
			self.vbox.addWidget(self.b)
			
			
			
		if bananas_to_cart > 0 and apples_to_cart[0] > 0 and grapes_to_cart == 0:
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-84599807.png')
			self.comboboxlabel211 = QLabel(self)
			self.comboboxlabel211.setPixmap(self.combobox21)
			self.comboboxlabel211.setGeometry(1, 1, 134, 89)
			self.lkp = QLabel(str(bananas_to_cart) + " pieces", self)
			self.lkp.showNormal()
			self.lkp.setGeometry(135, 1, 130, 99)
			self.lk21 = QLabel("$" + str(bananas_to_cart*70), self)
			self.lk21.showNormal()
			#self.lk2.setGeometry(135, 0, 130, 99)
			self.comboboxlabel211.showNormal()
			self.b90 = QPushButton(f"Buy {bananas_to_cart} bananas", self)
			self.b90.clicked.connect(self.btny)
			self.b90.setGeometry(1, 89, 134, 25)
			self.b90.setStyleSheet('background-color:green')
			self.b90.showNormal()
			try:
				if self.amount < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			except:
				if self.amoun < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel211)
			self.vbox.addWidget(self.lkp)
			self.vbox.addWidget(self.lk21)
			self.vbox.addWidget(self.b90)
			
			
			
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-78562683.png')
			self.comboboxlabel21 = QLabel(self)
			self.comboboxlabel21.setPixmap(self.combobox21)
			self.comboboxlabel21.setGeometry(1, 1+134, 134, 89)
			self.lk = QLabel(str(apples_to_cart[0]) + " pieces", self)
			self.lk.showNormal()
			self.lk.setGeometry(136, 137, 130, 99)
			self.lk2 = QLabel("$" + str(apples_to_cart[0]*50), self)
			self.lk2.setGeometry(1, 135, 134, 25)
			self.lk2.showNormal()
			self.comboboxlabel21.showNormal()
			self.b = QPushButton(f"Buy {apples_to_cart[0]} apples", self)
			self.b.clicked.connect(self.buy_me50)
			self.b.setGeometry(1, 135+89, 134, 25)
			self.b.setStyleSheet('background-color:green')
			self.b.showNormal()
			try:
				if self.amount < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			except:
				if self.amoun < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel21)
			self.vbox.addWidget(self.lk)
			self.vbox.addWidget(self.lk2)
			self.vbox.addWidget(self.b)
			
		if grapes_to_cart > 0 and apples_to_cart[0] == 0 and bananas_to_cart == 0:
			self.combobox2111 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-267191757.png')
			self.comboboxlabel2111 = QLabel(self)
			self.comboboxlabel2111.setPixmap(self.combobox2111)
			self.comboboxlabel2111.setGeometry(1, 1, 134, 89)
			self.lk111 = QLabel(str(grapes_to_cart) + " pieces", self)
			self.lk111.showNormal()
			self.lk111.setGeometry(135, 1, 130, 99)
			self.lk2111 = QLabel("$" + str(grapes_to_cart*30), self)
			self.lk2111.setGeometry(1, 1, 60, 25)
			self.lk2111.showNormal()
			self.comboboxlabel2111.showNormal()
			self.b111 = QPushButton(f"Buy {grapes_to_cart} grapes", self)
			self.b111.clicked.connect(self.buy_me11)
			self.b111.setGeometry(1, 89, 134, 25)
			self.b111.setStyleSheet('background-color:green')
			self.b111.showNormal()
			try:
				if self.amount < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			except:
				if self.amoun < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			
		if bananas_to_cart > 0 and apples_to_cart[0] == 0 and grapes_to_cart > 0:
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-84599807.png')
			self.comboboxlabel211 = QLabel(self)
			self.comboboxlabel211.setPixmap(self.combobox21)
			self.comboboxlabel211.setGeometry(1, 1, 134, 89)
			self.lkp = QLabel(str(bananas_to_cart) + " pieces", self)
			self.lkp.showNormal()
			self.lkp.setGeometry(135, 1, 130, 99)
			self.lk21 = QLabel("$" + str(bananas_to_cart*70))
			self.lk21.showNormal()
			#self.lk2.setGeometry(135, 0, 130, 99)
			self.comboboxlabel211.showNormal()
			self.b90 = QPushButton(f"Buy {bananas_to_cart} bananas", self)
			self.b90.clicked.connect(self.buy_me51111)
			self.b90.setGeometry(1, 89, 134, 25)
			self.b90.setStyleSheet('background-color:green')
			self.b90.showNormal()
			try:
				if self.amount < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			except:
				if self.amoun < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel211)
			self.vbox.addWidget(self.lkp)
			self.vbox.addWidget(self.lk21)
			self.vbox.addWidget(self.b90)
			
			
			self.combobox2111 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-267191757.png')
			self.comboboxlabel2111 = QLabel(self)
			self.comboboxlabel2111.setPixmap(self.combobox2111)
			self.comboboxlabel2111.setGeometry(1, 1+134, 134, 89)
			self.lk111 = QLabel(str(grapes_to_cart) + " pieces", self)
			self.lk111.showNormal()
			self.lk111.setGeometry(136, 137, 60, 25)
			self.lk2111 = QLabel("$" + str(grapes_to_cart*30), self)
			self.lk2111.setGeometry(1, 135, 134, 25)
			self.lk2111.showNormal()
			self.comboboxlabel2111.showNormal()
			self.b111 = QPushButton(f"Buy {grapes_to_cart} grapes", self)
			self.b111.clicked.connect(self.buy_me1110)
			self.b111.setGeometry(1, 135+89, 134, 25)
			self.b111.setStyleSheet('background-color:green')
			self.b111.showNormal()
			try:
				if self.amount < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			except:
				if self.amoun < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
					
			self.vbox.addWidget(self.comboboxlabel2111)
			self.vbox.addWidget(self.lk111)
			self.vbox.addWidget(self.lk2111)
			self.vbox.addWidget(self.b111)
			
			
		if apples_to_cart[0] > 0 and bananas_to_cart == 0 and grapes_to_cart == 0:
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-78562683.png')
			self.comboboxlabel21 = QLabel(self)
			self.comboboxlabel21.setPixmap(self.combobox21)
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk = QLabel(str(apples_to_cart[0]) + " pieces", self)
			self.lk.showNormal()
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2 = QLabel("$" + str(apples_to_cart[0]*50), self)
			self.lk2.showNormal()
			self.comboboxlabel21.showNormal()
			self.b = QPushButton(f"Buy {apples_to_cart[0]} apples", self)
			self.b.clicked.connect(self.buy_me50)
			self.b.setGeometry(1, 89, 134, 25)
			self.b.setStyleSheet('background-color:green')
			self.b.showNormal()
			try:
				if self.amount < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			except:
				if self.amoun < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel21)
			self.vbox.addWidget(self.lk)
			self.vbox.addWidget(self.lk2)
			self.vbox.addWidget(self.b)
			
		if grapes_to_cart > 0 and apples_to_cart[0] > 0 and bananas_to_cart == 0:
			self.combobox2111 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-267191757.png')
			self.comboboxlabel2111 = QLabel(self)
			self.comboboxlabel2111.setPixmap(self.combobox2111)
			self.comboboxlabel2111.setGeometry(1, 1, 134, 89)
			self.lk111 = QLabel(str(grapes_to_cart) + " pieces", self)
			self.lk111.showNormal()
			self.lk111.setGeometry(135, 1, 90, 25)
			self.lk2111 = QLabel("$" + str(grapes_to_cart*30), self)
			self.lk2111.setGeometry(1, 1, 60, 25)
			self.lk2111.showNormal()
			self.comboboxlabel2111.showNormal()
			self.b111 = QPushButton(f"Buy {grapes_to_cart} grapes", self)
			self.b111.clicked.connect(self.buy_me110)
			self.b111.setGeometry(1, 89, 134, 25)
			self.b111.setStyleSheet('background-color:green')
			self.b111.showNormal()
			try:
				if self.amount < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			except:
				if self.amoun < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
					
			self.vbox.addWidget(self.comboboxlabel2111)
			self.vbox.addWidget(self.lk111)
			self.vbox.addWidget(self.lk2111)
			self.vbox.addWidget(self.b111)
			
			
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-78562683.png')
			self.comboboxlabel21 = QLabel(self)
			self.comboboxlabel21.setPixmap(self.combobox21)
			self.comboboxlabel21.setGeometry(1, 1+134, 134, 89)
			self.lk = QLabel(str(apples_to_cart[0]) + " pieces", self)
			self.lk.showNormal()
			self.lk.setGeometry(136, 137, 60, 25)
			self.lk2 = QLabel("$" + str(apples_to_cart[0]*50), self)
			self.lk2.setGeometry(1, 135, 134, 25)
			self.lk2.showNormal()
			self.comboboxlabel21.showNormal()
			self.b = QPushButton(f"Buy {apples_to_cart[0]} apples", self)
			self.b.clicked.connect(self.buy_me511)
			self.b.setGeometry(1, 135+89, 134, 25)
			self.b.setStyleSheet('background-color:green')
			self.b.showNormal()
			try:
				if self.amount < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			except:
				if self.amoun < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			
			
			
		
		if bananas_to_cart > 0 and apples_to_cart[0] > 0 and grapes_to_cart > 0:
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-84599807.png')
			self.comboboxlabel211 = QLabel(self)
			self.comboboxlabel211.setPixmap(self.combobox21)
			self.comboboxlabel211.setGeometry(1, 1, 134, 89)
			self.lkp = QLabel(str(bananas_to_cart) + " pieces", self)
			self.lkp.showNormal()
			self.lkp.setGeometry(135, 1, 130, 99)
			self.lk21 = QLabel("$" + str(bananas_to_cart*70))
			self.lk21.showNormal()
			#self.lk2.setGeometry(135, 0, 130, 99)
			self.comboboxlabel211.showNormal()
			self.b90 = QPushButton(f"Buy {bananas_to_cart} bananas", self)
			self.b90.clicked.connect(self.buy_me51)
			self.b90.setGeometry(1, 89, 134, 25)
			self.b90.setStyleSheet('background-color:green')
			self.b90.showNormal()
			try:
				if self.amount < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			except:
				if self.amoun < bananas_to_cart*70:
					self.b90.setEnabled(False)
				else:
					self.b90.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel211)
			self.vbox.addWidget(self.lkp)
			self.vbox.addWidget(self.lk21)
			self.vbox.addWidget(self.b90)
			
			
			
			self.combobox21 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-78562683.png')
			self.comboboxlabel21 = QLabel(self)
			self.comboboxlabel21.setPixmap(self.combobox21)
			self.comboboxlabel21.setGeometry(1, 1+134, 134, 89)
			self.lk = QLabel(str(apples_to_cart[0]) + " pieces", self)
			self.lk.showNormal()
			self.lk.setGeometry(136, 137, 60, 25)
			self.lk2 = QLabel("$" + str(apples_to_cart[0]*50))
			self.lk2.setGeometry(1, 135, 134, 25)
			self.lk2.showNormal()
			self.comboboxlabel21.showNormal()
			self.b = QPushButton(f"Buy {apples_to_cart[0]} apples", self)
			self.b.clicked.connect(self.buy_me52)
			self.b.setGeometry(1, 135+89, 134, 25)
			self.b.setStyleSheet('background-color:green')
			self.b.showNormal()
			try:
				if self.amount < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			except:
				if self.amoun < apples_to_cart[0]*50:
					self.b.setEnabled(False)
				else:
					self.b.setEnabled(True)
			
			
			#self.hider2.showNormal()
			
			self.vbox.addWidget(self.comboboxlabel21)
			self.vbox.addWidget(self.lk)
			self.vbox.addWidget(self.lk2)
			self.vbox.addWidget(self.b)
			
			
			
			self.combobox2111 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-267191757.png')
			self.comboboxlabel2111 = QLabel(self)
			self.comboboxlabel2111.setPixmap(self.combobox2111)
			self.comboboxlabel2111.setGeometry(1, 1+134+91+25+50, 134, 89)
			self.lk111 = QLabel(str(grapes_to_cart) + " pieces", self)
			self.lk111.showNormal()
			self.lk111.setGeometry(136, 137+91+25+50, 90, 25)
			self.lk2111 = QLabel("$" + str(grapes_to_cart*30))
			self.lk2111.setGeometry(1, 1+134+91+25+50, 60, 25)
			self.lk2111.showNormal()
			self.comboboxlabel2111.showNormal()
			self.b111 = QPushButton(f"Buy {grapes_to_cart} grapes", self)
			#self.b111.setToolTip('Bought successfully')
			self.b111.clicked.connect(self.buy_me10)
			self.b111.setGeometry(1, 1+134+91+25+134, 134, 25)
			self.b111.setStyleSheet('background-color:green')
			self.b111.showNormal()
			try:
				if self.amount < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			except:
				if self.amoun < grapes_to_cart*30:
					self.b111.setEnabled(False)
				else:
					self.b111.setEnabled(True)
			
			self.vbox.addWidget(self.comboboxlabel211)
			self.vbox.addWidget(self.lkp)
			self.vbox.addWidget(self.lk21)
			self.vbox.addWidget(self.b90)
			
			self.vbox.addWidget(self.comboboxlabel21)
			self.vbox.addWidget(self.lk)
			self.vbox.addWidget(self.lk2)
			self.vbox.addWidget(self.b)
			
			self.vbox.addWidget(self.comboboxlabel2111)
			self.vbox.addWidget(self.lk111)
			self.vbox.addWidget(self.lk2111)
			self.vbox.addWidget(self.b111)
			
			
		self.hider = QPushButton("back", self)
		self.hider.setGeometry(1, 925, 60, 25)
		self.hider.setStyleSheet('background-color:black; color:white')
		self.hider.clicked.connect(self.back)
		self.hider.showNormal()
		self.vbox.addWidget(self.hider)
		
	def buy_me51111(self):
		self.b90.hide()
		self.lkp.hide()
		self.lk21.hide()
		self.comboboxlabel211.hide()
		
		try:
			global bananas_to_cart
		except:
			pass
			
		try:
			global grapes_to_cart
		except:
			pass
			
		try:
			self.amount -= bananas_to_cart*70
		except:
			self.amoun -= bananas_to_cart*70
			
		if grapes_to_cart == 0:
			self.label1man = QLabel( "❌", self)
			self.labeldouble = QLabel("No items in your cart", self)
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			bananas_to_cart -= bananas_to_cart
			
		if grapes_to_cart > 0:
			self.comboboxlabel2111.setGeometry(1, 1, 134, 89)
			self.lk2111.setGeometry(1, 1, 60, 25)
			self.lk111.setGeometry(135, 1, 130, 99)
			self.b111.setGeometry(1, 89, 134, 25)
			
			bananas_to_cart -= bananas_to_cart
	def buy_me110(self):
		self.comboboxlabel2111.hide()
		self.lk2111.hide()
		self.lk111.hide()
		self.b111.hide()
		try:
			global grapes_to_cart
		except:
			pass
		try:
			self.amount -= grapes_to_cart*30
		except:
			self.amoun -= grapes_to_cart*30
			
		if apples_to_cart[0] > 0:
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk.setGeometry(135, 1, 60, 25)
			self.lk2.setGeometry(1, 1, 60, 25)
			self.b.setGeometry(1, 89, 134, 25)
			grapes_to_cart -= grapes_to_cart
			
		if apples_to_cart[0] == 0:
			self.label1man = QLabel( "❌", self)
			self.labeldouble = QLabel("No items in your cart", self)
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			grapes_to_cart -= grapes_to_cart
		
	def buy_me1110(self):
		self.b111.hide()
		self.comboboxlabel2111.hide()
		self.lk111.hide()
		self.lk2111.hide()
		
		try:
			global grapes_to_cart
		except:
			pass
		try:
			self.amount -= grapes_to_cart*30
		except:
			self.amoun -= grapes_to_cart*30
		
			
		if bananas_to_cart == 0:
			self.label1man = QLabel( "❌", self)
			self.labeldouble = QLabel("No items in your cart", self)
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			
			
		grapes_to_cart -= grapes_to_cart
	def buy_me511(self):
		self.comboboxlabel21.hide()
		self.lk.hide()
		self.lk2.hide()
		self.b.hide()
		try:
			global apples_to_cart
		except:
			pass
		try:
			self.amount -= apples_to_cart[0]*50
		except:
			self.amoun -= apples_to_cart[0]*50
			
		if grapes_to_cart == 0:
			self.label1man = QLabel( "❌", self)
			self.labeldouble = QLabel("No items in your cart", self)
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			apples_to_cart[0] -= apples_to_cart[0]
			
		else:
			apples_to_cart[0] -= apples_to_cart[0]
			
	def buy_me11(self):
		self.comboboxlabel2111.hide()
		self.lk111.hide()
		self.lk2111.hide()
		self.b111.hide()
		try:
			global grapes_to_cart
		except:
			pass
		try:
			self.amount -= grapes_to_cart*30
		except:
			self.amoun -= grapes_to_cart*30
			
		self.label1man = QLabel( "❌", self)
		self.labeldouble = QLabel("No items in your cart", self)
		self.label1man.showNormal()
		self.labeldouble.showNormal()
		self.label1man.setGeometry(300, 400, 40, 30)
		self.labeldouble.setGeometry(250, 430, 150, 50)
		self.vbox.addWidget(self.label1man)
		self.vbox.addWidget(self.labeldouble)
		grapes_to_cart -= grapes_to_cart
			
			
		
		
	def buy_me10(self):
		self.comboboxlabel2111.hide()
		self.lk111.hide()
		self.lk2111.hide()
		self.b111.hide()
		try:
			self.amount -= grapes_to_cart*30
		except:
			self.amoun -= grapes_to_cart*30
			
	def buy_me51(self):
		self.comboboxlabel211.hide()
		self.lkp.hide()
		self.lk21.hide()
		self.b90.hide()
		try:
			self.amount -= bananas_to_cart*70
		except:
			self.amoun -= bananas_to_cart*70
			
		if grapes_to_cart > 0 and apples_to_cart[0] > 0:
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2.setGeometry(1, 1, 60, 25)
			self.b.setGeometry(1, 89, 134, 25)
			
			self.comboboxlabel2111.setGeometry(1, 1+134, 134, 89)
			self.lk111.setGeometry(136, 137, 60, 25)
			self.lk2111.setGeometry(136, 137, 60, 25)
			self.b111.setGeometry(1, 135+89, 134, 25)
			
			
			
		if grapes_to_cart > 0 and apples_to_cart[0] == 0:
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2.setGeometry(1, 1, 60, 25)
			self.b.setGeometry(1, 89, 134, 25)
		
		
	def buy_me52(self):
		self.comboboxlabel21.hide()
		self.lk.hide()
		self.lk2.hide()
		self.b.hide()
		try:
			self.amount -= apples_to_cart[0]*50
		except:
			self.amoun -= apples_to_cart[0]*50
			
		apples_to_cart[0] -= apples_to_cart[0]
			
#		if grapes_to_cart > 0 and bananas > 0
#			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
#			self.lk.setGeometry(135, 1, 130, 99)
#			self.lk2.setGeometry(1, 1, 60, 25)
#			self.b.setGeometry(1, 89, 134, 25)
#			
#			self.comboboxlabel2111.setGeometry(1, 1+134, 134, 89)
#			self.lk111.setGeometry(136, 137, 60, 25)
#			self.lk2111.setGeometry(136, 137, 60, 25)
#			self.b111.setGeometry(1, 135+89, 134, 25)
#			
#			
#			

		try:
			global bananas_to_cart
		except:
			pass
		if grapes_to_cart > 0 and bananas_to_cart == 0:
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2.setGeometry(1, 1, 60, 25)
			self.b.setGeometry(1, 89, 134, 25)
			
		if grapes_to_cart > 0 and bananas_to_cart > 0:
			self.comboboxlabel21.setGeometry(1, 1, 134, 89)
			self.lk.setGeometry(135, 1, 130, 99)
			self.lk2.setGeometry(1, 1, 60, 25)
			self.b.setGeometry(1, 89, 134, 25)
			
			self.comboboxlabel2111.setGeometry(1, 1+134, 134, 89)
			self.lk111.setGeometry(136, 137, 60, 25)
			self.lk2111.setGeometry(136, 137, 60, 25)
			self.b111.setGeometry(1, 135+89, 134, 25)
			
		if grapes_to_cart == 0 and bananas_to_cart == 0:
			self.label1man = QLabel( "❌")
			self.labeldouble = QLabel("No items in your cart")
			self.label1man.showNormal()
			self.labeldouble.showNormal()
			self.label1man.setGeometry(300, 400, 40, 30)
			self.labeldouble.setGeometry(250, 430, 150, 50)
			self.vbox.addWidget(self.label1man)
			self.vbox.addWidget(self.labeldouble)
			bananas_to_cart -= bananas_to_cart
	

		
		
		
		
		
			
			
	def makeplan(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		
		try:
			self.another_label.hide()
		except:
			pass
		
		try:
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.label1.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.hider.hide()
			
		except:
			pass
			
		try:
			self.label3.hide()
			self.button.hide()
			self.spin2.hide()
			self.hier.hide()
			
		except:
			pass
		
		self.btn6.clicked.connect(self.payin)
		self.b1 = QPushButton('Membership Plans', self)
		self.b1.setStyleSheet("background-color:red")
		
		self.gpbox2 = QGroupBox('Available plans: ')
		self.gpbox2.showNormal()
		
		grid2 = QGridLayout()
		grid2.addWidget(self.b1, 0,0)
		
		self.hider2 = QPushButton("back")
		self.hider2.setStyleSheet('background-color:black; color:white')
		self.hider2.showNormal()
		self.hider2.clicked.connect(self.back)
		
		self.b2 = QPushButton('Bundle Plans', self)
		self.b2.setStyleSheet('background-color:blue;')
		self.b2.showNormal()
		grid2.addWidget(self.b2, 0,1)
		
		self.b3 = QPushButton('Data Plans', self)
		self.b3.showNormal()
		self.b3.setStyleSheet("background-color:green")
		
		grid2.addWidget(self.b3, 1,0)
		grid2.addWidget(self.hider2, 1, 1)
		
		self.gpbox2.setLayout(grid2)
		
		self.vbox.addWidget(self.gpbox2)
		
	def payin(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		
		try:
			self.another_label.hide()
		except:
			pass
		
		try:
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.label1.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.hider.hide()
		except:
			pass
			
		try:
			self.b3.hide()
			self.b2.hide()
			self.gpbox2.hide()
			self.hider2.hide()
			self.b1.hide()
			
		except:
			pass
			
		try:
			self.comboboxlabel21.hide()
			self.lk.hide()
			self.lk2.hide()
			self.b.hide()
			self.hider2.hide()
		except:
			pass
			
		try:
			comboboxlabel211.hide()
			self.lkp.hide()
			self.lk21.hide()
			self.b90.hide()
		except:
			pass
		try:
			self.label1man.hide()
			self.labeldouble.hide()
		except:
			pass
			
		self.label3 = QLabel("Amount: ")
		self.label3.setGeometry(1, 100, 680, 500)
		self.label3.showNormal()
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(self.label3)
		self.button = QPushButton("Pay in")
		self.button.setStyleSheet('background-color:red; color:white')
		self.button.setGeometry(1, 500, 60, 25)
		self.button.showNormal()
		self.button.clicked.connect(self.pay)
		self.spin2 = QSpinBox()
		self.spin2.setGeometry(1, 400, 600, 25)
		self.spin2.setMaximum(10000)
		self.spin2.showNormal()
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(self.spin2)
		self.vbox.addWidget(self.button, alignment=Qt.AlignLeft)
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		self.hier = QPushButton("back", self)
		self.hier.setStyleSheet('background-color:black; color:white')
		self.hier.showNormal()
		self.hier.clicked.connect(self.back)
		self.vbox.addWidget(self.hier, alignment=Qt.AlignLeft)
		
	def pay(self):
		payer = self.spin2.value()
		message = QMessageBox.question(self, 'Payment', 'Payed in successfully', QMessageBox.Ok)
		if message == QMessageBox.Ok:
			try:
				self.amount += payer
			except:
				self.amoun += payer
		else:
			#raise Exception(dir(QMessageBox))
			m = QMessageBox.alert(self, 'Cancelled', 'Payment cancelled')
			#m = QMessageBox.Question
				
				
				
				
				
				
	def buy(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		
		try:
			self.another_label.hide()
		except:
			pass
		
		try:
			self.label3.hide()
			self.spin2.hide()
			self.button.hide()
			self.hier.hide()
			
#			self.vbox.removeWidget(self.label3)
#			self.vbox.removeWidget(self.button)
#			self.vbox.removeWidget(self.spin2)
		except:
			pass
			
		try:
			self.b3.hide()
			self.b2.hide()
			self.gpbox2.hide()
			self.hider2.hide()
			self.b1.hide()
		except:
			pass
			
		try:
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.label1.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.comboboxlabel2.hide()
			self.btn194.hide()
			self.btn195.hide()
		except:
			pass
			
		try:
			self.comboboxlabel21.hide()
			self.lk.hide()
			self.lk2.hide()
			self.b.hide()
		except:
			pass
			
		try:
			self.comboboxlabel211.hide()
			self.lkp.hide()
			self.lk21.hide()
			self.b90.hide()
		except:
			pass
			
		try:
			self.label1man.hide()
			self.labeldouble.hide()
		except:
			pass
			
		self.combobox = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-78562683.png')
		self.comboboxlabel = QLabel(self)
		self.comboboxlabel.setPixmap(self.combobox)
		
		self.combobox1 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-84599807.png')
		self.comboboxlabel1 = QLabel(self)
		self.comboboxlabel2 = QLabel(self)
		self.combobox2 = QPixmap('/storage/emulated/0/Download/Attachments/Screenshot_20200712-151452_kindlephoto-267191757.png')
		self.comboboxlabel2.setPixmap(self.combobox2)
		self.comboboxlabel2.showNormal()
		self.comboboxlabel1.setPixmap(self.combobox1)
		self.btn190 = QPushButton("Buy")
		self.btn190.setStyleSheet("color:white; background-color:green;")
		self.btn190.clicked.connect(self.buy_apples)
		self.btn191 = QPushButton("Add to cart")
		self.btn191.setStyleSheet("color:white; background-color:green;")
		self.btn191.clicked.connect(self.cart_apples)
		
		self.btn192 = QPushButton("Buy")
		self.btn192.setStyleSheet("color:white; background-color:green;")
		self.btn192.clicked.connect(self.buy_bananas)
		self.btn193 = QPushButton("Add to cart")
		self.btn193.setStyleSheet("color:white; background-color:green;")
		self.btn193.clicked.connect(self.cart_bananas)
		self.btn194 = QPushButton( "Buy")
		self.btn194.setStyleSheet("background-color:green; color:white")
		self.btn194.clicked.connect(self.buy_grapes)
		self.btn194.showNormal()
		self.btn195 = QPushButton("Add to cart")
		self.btn195.setStyleSheet("background-color:green; color:white")
		self.btn195.clicked.connect(self.cart_grapes)
		self.btn195.showNormal()
		
		self.label1 = QLabel("Product: ")
		self.label1.showNormal()
		self.comboboxlabel.setGeometry(1, 100, 134, 89)
		self.comboboxlabel.showNormal()
		self.btn190.setGeometry(1, 200, 60, 25)
		self.btn190.showNormal()
		self.btn191.setGeometry(1, 225, 90, 25)
		self.btn191.showNormal()
		self.comboboxlabel1.setGeometry(1, 290, 134, 89)
		self.comboboxlabel1.showNormal()
		self.btn192.setGeometry(1, 390, 60, 25)
		self.btn192.showNormal()
		self.btn193.setGeometry(1, 390+25, 90, 25)
		self.btn193.showNormal()
		self.comboboxlabel2.setGeometry(1, 390+25+65, 134, 89)
		self.comboboxlabel2.showNormal()
		self.btn194.setGeometry(1, 390+25+65+100, 60, 25)
		self.btn194.showNormal()
		self.btn195.setGeometry(1, 390+25+65+100+25, 90, 25)
		self.btn195.showNormal()
		
		self.gbtn = QPushButton("Buy material")
		self.gbtn.setStyleSheet('background-color:red')
		#self.gbtn.clicked.connect(self.buy_your_material)
		#hbox = QHBoxLayout()
		#hbox.addWidget(self.comboboxlabel)
		self.hider = QPushButton("back")
		#self.hider.setGeometry(1, 925, 60, 25)
		self.hider.setStyleSheet('background-color:black;color:white')
		self.hider.clicked.connect(self.back)

		
		#self.hider.setGeometry(100, 100, 50, 50)
		self.vbox.addWidget(self.label1)
		self.vbox.addWidget(self.comboboxlabel)
		self.vbox.addWidget(self.btn190, alignment=Qt.AlignLeft)
		self.vbox.addWidget(self.btn191, alignment=Qt.AlignLeft)
		#self.vbox.addWidget(QLabel(""))
		#raise Exception(dir(Qt))
		self.vbox.addWidget(self.comboboxlabel1)
		self.vbox.addWidget(self.btn192, alignment=Qt.AlignLeft)
		self.vbox.addWidget(self.btn193, alignment=Qt.AlignLeft)
		
		self.vbox.addWidget(self.comboboxlabel2)
		self.vbox.addWidget(self.btn194, alignment=Qt.AlignLeft)
		self.vbox.addWidget(self.btn195, alignment=Qt.AlignLeft)
		#self.vbox.addWidget(self.label2
		#self.vbox.addWidget(self.box)
		self.vbox.addWidget(QLabel(""))
		self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(hbox)
		#self.vbox.addWidget(self.gbtn)
		self.hider.showNormal()
		self.vbox.addWidget(self.hider)
		#self.vbox.addWidget(self.hider)
		
	def buy_apples(self):
		message = QInputDialog.getInt(self, "Quantity", 'How many apples do you want to buy')
		me = QMessageBox.question(self, "Ensure", f"You are just about to add {message[0]} apples to cart, each has $50 that will cost you ${50*message[0]}")
		if me == QMessageBox.Yes:
			try:
				if self.amount < 50*message[0]:
					m = QMessageBox.question(self, "Error", 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					global apples
					apples[0] += message[0]
					self.amount -= 50*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
			except:
				if self.amoun < 50*message[0]:
					m = QMessageBox.question(self, "Error", 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					#global apples
					apples[0] += message[0]
					self.amoun -= 50*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
					#self.back()
				
	def buy_bananas(self):
		message = QInputDialog.getInt(self, "Quantity", 'How many bananas do you want to buy')
		me = QMessageBox.question(self, "Ensure", f"You are just about to buy {message[0]} bananas and each costs $70 and will remove ${70*message[0]}")
		if me == QMessageBox.Yes:
			try:
				if self.amount < 70*message[0]:
					m = QMessageBox.question(self, 'Error', 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					global bananas
					bananas += message[0]
					self.amount -= 70*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
			except:
				if self.amoun < 70*message[0]:
					m = QMessageBox.question(self, "Error", 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					bananas += message[0]
					self.amoun -= 70*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
					
	def buy_grapes(self):
		message = QInputDialog.getInt(self, "Quantity", 'How many grapes do you want to buy')
		me = QMessageBox.question(self, "Ensure", f"You are just about to buy {message[0]} grapes and each costs $30 and will remove ${30*message[0]}")
		if me == QMessageBox.Yes:
			try:
				if self.amount < 30*message[0]:
					m = QMessageBox.question(self, 'Error', 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					global grapes
					grapes += message[0]
					self.amount -= 30*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
			except:
				if self.amoun < 30*message[0]:
					m = QMessageBox.question(self, "Error", 'Item bought unsuccessfully', QMessageBox.Ok)
				else:
					grapes += message[0]
					self.amoun -= 30*message[0]
					m = QMessageBox.about(self, 'Success', 'Item bought successfully')
					
					
	def cart_bananas(self):
		#self.label2 = QLabel("Amount: ")
		#self.box = QSpinBox()
	#	self.gbtn.clicked.connect(self.cart_it)
		#self.vbox.removeWidget(self.gbtn)
		self.vbox.removeWidget(self.hider)
#		self.vbox.addWidget(self.label2)
#		self.vbox.addWidget(self.box)
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(self.gbtn)
#		self.vbox.addWidget(self.hider)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
#		self.vbox.removeWidget(self.comboboxlabel)
#		self.vbox.removeWidget(self.comboboxlabel1)
#		self.vbox.removeWidget(self.btn190)
#		self.vbox.removeWidget(self.btn191)
#		self.vbox.removeWidget(self.btn192)
#		self.vbox.removeWidget(self.btn193)
#		self.vbox.removeWidget(self.label1)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
		message = QInputDialog.getInt(self, "Quantity", 'How many bananas do you want to add to cart')
		me = QMessageBox.question(self, "Ensure", f"You are just about to add {message[0]} bananas to cart")
		if me == QMessageBox.Yes:
			global bananas_to_cart
			bananas_to_cart += message[0]
			m = QMessageBox.about(self, 'Success', 'Item added to cart successfully')
			
	def cart_grapes(self):
		#self.label2 = QLabel("Amount: ")
		#self.box = QSpinBox()
	#	self.gbtn.clicked.connect(self.cart_it)
		#self.vbox.removeWidget(self.gbtn)
		#self.vbox.removeWidget(self.hider)
#		self.vbox.addWidget(self.label2)
#		self.vbox.addWidget(self.box)
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(self.gbtn)
#		self.vbox.addWidget(self.hider)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
#		self.vbox.removeWidget(self.comboboxlabel)
#		self.vbox.removeWidget(self.comboboxlabel1)
#		self.vbox.removeWidget(self.btn190)
#		self.vbox.removeWidget(self.btn191)
#		self.vbox.removeWidget(self.btn192)
#		self.vbox.removeWidget(self.btn193)
#		self.vbox.removeWidget(self.label1)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
		message = QInputDialog.getInt(self, "Quantity", 'How many grapes do you want to add to cart')
		me = QMessageBox.question(self, "Ensure", f"You are just about to add {message[0]} grapes to cart")
		if me == QMessageBox.Yes:
			global grapes_to_cart
			grapes_to_cart += message[0]
			m = QMessageBox.about(self, 'Success', 'Item added to cart successfully')
			
			
			
	def cart_apples(self):
		#self.label2 = QLabel("Amount: ")
		#self.box = QSpinBox()
	#	self.gbtn.clicked.connect(self.cart_it)
		#self.vbox.removeWidget(self.gbtn)
#		self.vbox.removeWidget(self.hider)
#		self.vbox.addWidget(self.label2)
#		self.vbox.addWidget(self.box)
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(self.gbtn)
#		self.vbox.addWidget(self.hider)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
#		self.vbox.removeWidget(self.comboboxlabel)
#		self.vbox.removeWidget(self.comboboxlabel1)
#		self.vbox.removeWidget(self.btn190)
#		self.vbox.removeWidget(self.btn191)
#		self.vbox.removeWidget(self.btn192)
#		self.vbox.removeWidget(self.btn193)
#		self.vbox.removeWidget(self.label1)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
		message = QInputDialog.getInt(self, "Quantity", 'How many apples do you want to add to cart')
		me = QMessageBox.question(self, "Ensure", f"You are just about to add {message[0]} apples to cart")
		if me == QMessageBox.Yes:
			global apples_to_cart
			apples_to_cart[0] += message[0]
			m = QMessageBox.about(self, 'Success', 'Item added to cart successfully')
		
		
	def cart_bananas(self):
		#self.label2 = QLabel("Amount: ")
		#self.box = QSpinBox()
	#	self.gbtn.clicked.connect(self.cart_it)
		#self.vbox.removeWidget(self.gbtn)
		#self.vbox.removeWidget(self.hider)
#		self.vbox.addWidget(self.label2)
#		self.vbox.addWidget(self.box)
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(QLabel(""))
		#self.vbox.addWidget(self.gbtn)
#		self.vbox.addWidget(self.hider)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
#		self.vbox.removeWidget(self.comboboxlabel)
#		self.vbox.removeWidget(self.comboboxlabel1)
#		self.vbox.removeWidget(self.btn190)
#		self.vbox.removeWidget(self.btn191)
#		self.vbox.removeWidget(self.btn192)
#		self.vbox.removeWidget(self.btn193)
#		self.vbox.removeWidget(self.label1)
#		self.comboboxlabel1.hide()
#		self.comboboxlabel.hide()
#		self.btn190.hide()
#		self.btn191.hide()
#		self.btn192.hide()
#		self.btn193.hide()
#		self.label1.hide()
		message = QInputDialog.getInt(self, "Quantity", 'How many bananas do you want to add to cart')
		me = QMessageBox.question(self, "Ensure", f"You are just about to add {message[0]} bananas to cart")
		if me == QMessageBox.Yes:
			global bananas_to_cart
			bananas_to_cart += message[0]
			m = QMessageBox.about(self, 'Success', 'Item added to cart successfully')
	def back(self):
		#Remove all elements()
		#self.box.hide()
		#self.gbtn.hide()
		try:
			self.label1.hide()
			self.comboboxlabel.hide()
			self.comboboxlabel1.hide()
			self.comboboxlabel2.hide()
			self.btn190.hide()
			self.btn191.hide()
			self.btn192.hide()
			self.btn193.hide()
			self.btn194.hide()
			self.btn195.hide()
			#self.btn5.show()
		except:
			pass
			
		#self.label2.hide()
		#self.hider.hide()
		#raise Exception(dir(self.comboboxlabel))
		#self.comboboxlabel.hide()
		self.load()
		
		#Show first page
		#try:
#		self.btn5.show()
#		self.btn6.show()
#		self.btn.show()
#		self.btn4.show()
#		self.btn3.show()
#		self.btn2.show()
#		self.gpbox.show()
#		self.btn1.show()
	def detail(self):
		try:
			self.btn5.hide()
			self.btn6.hide()
			self.btn.hide()
			self.btn4.hide()
			self.btn3.hide()
			self.btn2.hide()
			self.gpbox.hide()
			self.btn1.hide()
			self.j
			message = QMessageBox.about(self, 'Details', f'Username: {self.check1}\nEmail: {self.email}\nDate of registry: {self.reg}\nUser Balance: ${self.amount}\nUser Data: {self.data}')
			self.btn5.show()
			self.btn6.show()
			self.btn.show()
			self.btn4.show()
			self.btn3.show()
			self.btn2.show()
			self.gpbox.show()
			self.btn1.show()
		except:
			message = QMessageBox.about(self, 'Error', 'You Need to login before finding details')
			self.btn5.show()
			self.btn6.show()
			self.btn.show()
			self.btn4.show()
			self.btn3.show()
			self.btn2.show()
			self.btn1.show()
			self.gpbox.show()
		
	def signin(self):
		self.btn5.hide()
		self.btn6.hide()
		self.btn.hide()
		self.btn4.hide()
		self.btn3.hide()
		self.btn2.hide()
		self.gpbox.hide()
		self.btn1.hide()
		self.check1, done1 = QInputDialog.getText(self, 'Username', 'Enter your name:')
		wb = xl.load_workbook('/storage/emulated/0/Documents/UserData.xlsx')
		sheet = wb["Users"]
		names = []
		for name in range(wb["Data_space"]["B4"].value):
			names.append(sheet["B"+str(name+2)])
		QMessageBox.information(self, '-', str(names))
#		if str(self.check1) == "Ozichi":
#			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/First Customer Information.xlsx")
#			sheet = wb["First Customer Information.csv"]
#			name, done1 = QInputDialog.getText(self, 'Security', 'Enter your password:')
#			if str(name) != sheet["B5"].value:
#		 		message = QMessageBox.about(self, 'Error', 'Wrong password')
#		 		self.btn5.show()
#			 	self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.gpbox.show()
#			else:
#		 		message = QMessageBox.question(self, 'Correct', f'Welcome back Ozichi\n{sheet["B9"].value}', QMessageBox.Ok)
#		 		self.btn5.show()
#		 		self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.gpbox.show()
#		 		self.j = "8"
#		 		self.email = sheet['C4'].value
#		 		self.reg = sheet['D4'].value
#		 		self.amount = int(sheet['B2'].value)
#		 		self.data = int(sheet['B3'].value)
#		 		login = True
#		elif str(self.check1) == "ifechi":
#			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/Fourth customer information.xlsx")
#			sheet = wb["Sheet1"]
#			name, done1 = QInputDialog.getText(self, 'Security', 'Enter your password:')
#			if str(name) != sheet["C5"].value:
#		 		message = QMessageBox.about(self, 'Error', 'Wrong password')
#		 		self.btn5.show()
#			 	self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.gpbox.show()
#			else:
#		 		message = QMessageBox.question(self, 'Correct', f'Welcome back ifechi\n{sheet["B9"].value}', QMessageBox.Ok)
#		 		self.btn5.show()
#		 		self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.j = "8"
#		 		self.email = sheet['C4'].value
#		 		self.reg = sheet['D4'].value
#		 		self.amount = int(sheet['B2'].value)
#		 		self.data = int(sheet['B3'].value)
#		 		self.gpbox.show()
#		 		login = True
#		elif str(self.check1) == "Akachi":
#			wb = xl.load_workbook("/storage/emulated/0/Lucky bundle cache/Third customer Information.xlsx")
#			sheet = wb["Third customer Information.csv"]
#			name, done1 = QInputDialog.getText(self, 'Security', 'Enter your password:')
#			if str(name) != sheet["B5"].value:
#		 		message = QMessageBox.about(self, 'Error', 'Wrong password')
#		 		self.btn5.show()
#			 	self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.gpbox.show()
#			else:
#		 		message = QMessageBox.question(self, 'Correct', f'Welcome back Akachi\n{sheet["B9"].value}', QMessageBox.Ok)
#		 		self.btn5.show()
#		 		self.btn6.show()
#		 		self.btn.show()
#		 		self.btn4.show()
#		 		self.btn3.show()
#		 		self.btn2.show()
#		 		self.btn1.show()
#		 		self.gpbox.show()
#		 		self.j = "8"
#		 		self.email = sheet['C4'].value
#		 		self.reg = sheet['D4'].value
#		 		self.amount = int(sheet['B2'].value)
#		 		self.data = int(sheet['B3'].value)
#		 		login = True
#		else:
#		 	message = QMessageBox.question(self, 'Error', 'No such account', QMessageBox.Ok)
#		 	self.btn5.show()
#		 	self.btn6.show()
#		 	self.btn.show()
#		 	self.btn4.show()
#		 	self.btn3.show()
#		 	self.btn2.show()
#		 	self.btn1.show()
#		 	self.gpbox.show()
if __name__ == "__main__":
	app = QApplication(sys.argv)
	window = Window()
	sys.exit(app.exec_())